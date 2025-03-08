"""
سيرفر Python لمعالجة ملفات Excel مع المحافظة على التنسيق الأصلي
====================================================
هذا السيرفر يستخدم Flask لإنشاء واجهة برمجة تطبيقات RESTful تسمح بتحميل ملفات Excel،
إضافة التعليقات والتقديرات، والحفاظ على التنسيق الأصلي للملف.

المتطلبات: 
- Python 3.6+
- Flask
- openpyxl
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
import openpyxl
import tempfile
import json
import logging
from datetime import datetime

# إعداد السجلات
logging.basicConfig(level=logging.INFO,
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)  # تمكين CORS للسماح بالطلبات من المتصفح

# مجلد مؤقت لتخزين الملفات
TEMP_FOLDER = tempfile.gettempdir()

@app.route('/api/status', methods=['GET'])
def check_status():
    """التحقق من حالة الخادم"""
    return jsonify({
        "status": "online",
        "message": "خادم Python متصل ويعمل",
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })

@app.route('/api/process-excel', methods=['POST'])
def process_excel():
    """معالجة ملف Excel وإضافة التقديرات مع الحفاظ على التنسيق الأصلي"""
    try:
        # التحقق من وجود الملف
        if 'file' not in request.files:
            return jsonify({"error": "لم يتم تحميل ملف"}), 400

        excel_file = request.files['file']
        student_data_json = request.form.get('studentData')
        sheet_name = request.form.get('sheetName')

        if not excel_file or not student_data_json or not sheet_name:
            return jsonify({"error": "بيانات غير مكتملة"}), 400

        # تحليل بيانات الطلاب من JSON
        student_data = json.loads(student_data_json)

        # إنشاء اسم ملف فريد
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_path = os.path.join(TEMP_FOLDER, f"temp_excel_{timestamp}.xlsx")
        output_path = os.path.join(TEMP_FOLDER, f"processed_excel_{timestamp}.xlsx")

        # حفظ الملف المستلم
        excel_file.save(temp_path)
        logger.info(f"تم حفظ الملف المؤقت: {temp_path}")

        # فتح الملف باستخدام openpyxl للحفاظ على التنسيق
        workbook = openpyxl.load_workbook(temp_path)
        
        # التحقق من وجود الورقة المطلوبة
        if sheet_name not in workbook.sheetnames:
            return jsonify({"error": f"لم يتم العثور على الورقة: {sheet_name}"}), 400
        
        worksheet = workbook[sheet_name]
        logger.info(f"تم فتح الورقة: {sheet_name}")

        # إضافة التقديرات إلى العمود H
        for student in student_data:
            # الحصول على الصف المناسب من بيانات الطالب
            row_index = student.get('rowIndex')
            if not row_index:
                continue
            
            # الحصول على التقدير
            grade_comment = student.get('gradeComment', '')
            
            # إضافة التقدير إلى العمود H في الصف المحدد
            cell = worksheet.cell(row=row_index, column=8)  # العمود H هو رقم 8
            cell.value = grade_comment
            
            logger.info(f"تمت إضافة التقدير '{grade_comment}' في الخلية H{row_index}")

        # حفظ الملف المعدل
        workbook.save(output_path)
        logger.info(f"تم حفظ الملف المعالج: {output_path}")

        # إرسال الملف للتحميل
        return send_file(
            output_path,
            as_attachment=True,
            download_name=f"نتائج_{sheet_name}_{timestamp}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"حدث خطأ: {str(e)}")
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    logger.info("بدء تشغيل خادم تحليل ملفات Excel...")
    print("تم بدء تشغيل خادم Python على المنفذ 5000")
    print("استخدم Ctrl+C لإيقاف الخادم")
    # تشغيل التطبيق على المنفذ 5000 ومتاح من أي عنوان IP
    app.run(host='0.0.0.0', port=5000, debug=True)
